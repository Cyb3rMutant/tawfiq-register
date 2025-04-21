from flask import Flask, json, request, render_template, url_for, redirect, session
from model import model
from datetime import datetime, timedelta
import io
import openpyxl

# from passlib.hash import sha256_crypt

app = Flask(__name__)
app.secret_key = "a1V#9R^l6vhGI'Xyms@]ARPJ"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=30)
server_last_update_time = datetime.now()


@app.route("/")
def index():
    if "teacher_id" not in session:
        return redirect(url_for("login"))

    classes = model.get_teacher_classes(teacher_id=session.get("teacher_id"))
    return render_template("index.html", classes=classes)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method != "POST":
        return render_template("login.html")

    username = request.form["username"]
    user = model.get_teacher(username)
    print(user)
    if not user:
        return render_template("login.html", message="user doesn't exist")

    session["teacher_id"] = user["teacher_id"]

    return redirect(url_for("index"))  # Redirect after adding the class


@app.route("/add_student", methods=["GET", "POST"])
def add_student():
    if request.method != "POST":
        payers = model.get_payers()
        classes = model.get_classes()
        return render_template("add_student.html", classes=classes, payers=payers)

    full_name = request.form["full_name"]
    age = int(request.form["age"])
    gender = request.form["gender"]
    phone_number = request.form["phone_number"]
    classes = request.form.getlist("classes")

    # Handle payer selection or new payer creation
    payer_id = request.form["payer_ids"]
    if payer_id == "-1":  # New payer creation
        new_payer_name = request.form["new_payer_name"]
        new_payer_phone = request.form["new_payer_phone"]
        payer_id = model.add_payer(
            new_payer_name, new_payer_phone
        )  # Function to insert new payer
    elif payer_id == "0":
        payer_id = model.add_payer(full_name, phone_number)

    # Add student and associate with payer
    model.add_student(full_name, age, gender, phone_number, classes, payer_id)

    return redirect(url_for("index"))


@app.route("/add_teacher", methods=["GET", "POST"])
def add_teacher():
    if request.method != "POST":
        # Render the form
        return render_template("add_teacher.html")
    teacher_name = request.form["teacher_name"]
    phone_number = request.form["phone_number"]

    # Call the add_teacher function
    model.add_teacher(teacher_name=teacher_name, phone_number=phone_number)

    return redirect(url_for("index"))  # Redirect after adding teacher


def parse_fields(fields):
    field_types = model.get_field_types()
    fields = json.loads(fields)
    for field in fields:
        for t in field_types:
            if int(field["field_type_id"]) == t["field_type_id"]:
                if not t["field_type_defaults"]:
                    field["defaultValues"] = []
        else:
            field["defaultValues"] = json.dumps(field["defaultValues"])
    return fields


@app.route("/add_class", methods=["GET", "POST"])
def add_class():
    field_types = model.get_field_types()
    days_of_week = [
        "Monday",
        "Tuesday",
        "Wednesday",
        "Thursday",
        "Friday",
        "Saturday",
        "Sunday",
    ]
    if request.method != "POST":
        # For GET request: Retrieve available teachers for the form
        teachers = model.get_teachers()
        packages = model.get_packages()
        return render_template(
            "add_class.html",
            teachers=teachers,
            packages=packages,
            days_of_week=days_of_week,
            field_types=field_types,
        )

    class_name = request.form["class_name"]
    teacher_ids = request.form.getlist("teacher_ids")  # Multiple teachers
    package_id = request.form["package_id"]
    selected_days = request.form.getlist("days_of_week")  # Get list of selected days
    days_binary = "".join("1" if day in selected_days else "0" for day in days_of_week)
    time_of_day = request.form["time_of_day"]
    fields = parse_fields(request.form["fields"])

    # Call the function to add a class
    model.add_class(
        class_name=class_name,
        teacher_ids=teacher_ids,
        package_id=package_id,
        days_of_week=days_binary,
        time_of_day=time_of_day,
        fields=fields,
    )

    return redirect(url_for("index"))  # Redirect after adding the class


@app.route("/add_package", methods=["GET", "POST"])
def add_package():
    if request.method != "POST":
        return render_template("add_package.html")

    package_name = request.form["package_name"]
    price = request.form["price"]

    # Call the function to add a package
    model.add_package(
        package_name=package_name,
        price=price,
    )

    return redirect(url_for("index"))  # Redirect after adding the class


@app.route("/assign_students", methods=["GET", "POST"])
def assign_students():
    if request.method != "POST":
        # For GET request: Retrieve available students and classes
        students = model.get_students()

        classes = model.get_classes()

        # Render the form
        return render_template(
            "assign_students.html", students=students, classes=classes
        )

    class_id = request.form["class_id"]
    student_ids = request.form.getlist("student_ids")  # Get multiple selected students

    # Call the function to assign students
    model.assign_students_to_class(class_id=class_id, student_ids=student_ids)

    return redirect(url_for("index"))  # Redirect after adding the class


@app.route("/ajax_server/", methods=["POST", "GET"])
def ajax_server():
    print("called")
    global server_last_update_time
    if request.method == "GET":
        field_id = request.args.get("field_id")
        field = request.args.get("field")
        value = request.args.get("value")
        print(field_id, field, value)
        model.update_attendance_field(field_id, field, value)

        server_last_update_time = datetime.now()
        return str(server_last_update_time)


@app.route("/ajax_poll", methods=["POST", "GET"])
def ajax_poll():
    global server_last_update_time
    if request.method == "GET":
        client_last_update_time = datetime.strptime(
            request.args.get("value"), "%d/%m/%Y %H:%M:%S"
        )
        if client_last_update_time > server_last_update_time:
            return []
        attendance_date = datetime.now() + timedelta(days=7 * 0)
        class_id = request.args.get("class_id")
        return model.get_attendance(class_id, attendance_date)


def decode_days_binary(binary_string):
    # Days of the week in order: Monday, Tuesday, Wednesday, Thursday, Friday, Saturday, Sunday
    days_of_week = [
        "Monday",
        "Tuesday",
        "Wednesday",
        "Thursday",
        "Friday",
        "Saturday",
        "Sunday",
    ]

    # Decode the binary string into days that are selected
    selected_days = [days_of_week[i] for i in range(7) if binary_string[i] == "1"]

    return selected_days


@app.route("/mark_attendance/<int:class_id>", methods=["GET", "POST"])
def mark_attendance(class_id):
    # Get the current date and day of the week
    attendance_date = datetime.now()
    current_day = attendance_date.strftime("%A")  # Get current day (e.g., "Monday")

    # Fetch the class data from the database using the class_id
    class_data = model.get_class(class_id)
    print(
        f"Attendance Date: {attendance_date}, Current Day: {current_day}, Class Days: {class_data['days_of_week']}"
    )

    # Decode the binary days string into a list of active days
    active_days = decode_days_binary(class_data["days_of_week"])
    print(f"Active Days for Class {class_data['class_name']}: {active_days}")

    # Check if the class is running today
    if current_day not in active_days:
        return f"{class_data['class_name']} is not running today."
    if not model.get_class_payments(class_id, attendance_date):
        model.init_payment_month(class_id, attendance_date)
    attendance = model.get_attendance(class_id, attendance_date)
    if not attendance:
        # mark all students as none attendant
        model.init_attendance_day(class_id, attendance_date)
        attendance = model.get_attendance(class_id, attendance_date)

    print(attendance)
    if request.method != "POST":
        return render_template(
            "mark_attendance.html", attendance=attendance, class_data=class_data
        )

    # note just collect attendance separatly then payments separatly then fields separatly since they have their own ids no need to be bound to student id
    attendance_vals = {}
    payment_vals = {}
    field_vals = {}

    for data in attendance:
        attendance_vals[data["attendance_id"]] = request.form[
            "student_{}".format(data["attendance_id"])
        ]
        payment_vals[data["payment_id"]] = int(
            request.form.get("paid_{}".format(data["payment_id"]), "off") == "on"
        )

        for field in data["fields"]:

            field_values = request.form.getlist(
                "field_{}".format(field["attendance_field_id"])
            )  # Handles checkbox
            # Convert values to integers if possible
            try:
                field_values = [int(v) for v in field_values]
            except ValueError:
                pass  # Keep them as strings if conversion fails
            field_vals[field["attendance_field_id"]] = field_values

    print(attendance_vals, payment_vals, field_vals)
    model.mark_attendance(attendance_vals, payment_vals, field_vals)
    return redirect(url_for("index"))  # Redirect after adding the class


from datetime import datetime, timedelta


def parse_fields_2(data):
    for entry in data:
        parsed_entry = {"status": entry["status"]}
        for field in entry["fields"]:
            field_name = field.get("field_name")
            field_value = field.get("field_value")
            field_type_name = field.get("field_type_name")
            field_defaults = field.get("field_defaults", [])

            if not field_value:
                parsed_entry[field_name] = ""
                continue

            # Handle text field
            if field_type_name == "text":
                parsed_entry[field_name] = field_value[0]

            # Handle radio field
            elif field_type_name == "radio":
                selected_index = field_value[0]
                # Get the corresponding label from field_defaults
                selected_label = next(
                    (
                        label
                        for value, label in field_defaults
                        if value == selected_index
                    ),
                    None,
                )
                parsed_entry[field_name] = selected_label

            # Handle checkbox field
            elif field_type_name == "checkbox":
                selected_labels = [
                    label for value, label in field_defaults if value in field_value
                ]
                parsed_entry[field_name] = selected_labels

        entry["fields"] = parsed_entry


@app.route("/view_attendance/<int:class_id>")
def view_attendance(class_id):
    from pprint import pprint

    # Get the start_date from the query string or default to the current date
    start_date_str = request.args.get("start_date", None)
    if start_date_str:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    else:
        start_date = datetime.today().replace(day=1)  # First day of the current month

    # Calculate the end_date (3 months from start_date)
    end_date = start_date + timedelta(days=30)  # Approx. 3 months

    # Fetch class data and attendance filtered by date range
    class_data = model.get_class(class_id)
    attendance = model.get_class_attendance_in_date_range(
        class_id, start_date, end_date
    )
    parse_fields_2(attendance)

    # Extract unique names and dates within the range
    names = sorted(set(item["full_name"] for item in attendance))
    dates = sorted(set(item["date"] for item in attendance))

    # Build a lookup dictionary for quick access
    lookup = {(item["full_name"], item["date"]): item["fields"] for item in attendance}
    all_fields = list(lookup.values())[0].keys() if lookup else [""]

    # Calculate previous and next start dates
    prev_start_date = (start_date - timedelta(days=30)).strftime("%Y-%m-%d")
    next_start_date = (start_date + timedelta(days=30)).strftime("%Y-%m-%d")

    return render_template(
        "view_attendance.html",
        names=names,
        dates=dates,
        lookup=lookup,
        all_fields=all_fields,
        class_data=class_data,
        prev_start_date=prev_start_date,
        next_start_date=next_start_date,
        current_period=f"{start_date.strftime('%b %Y')} - {end_date.strftime('%b %Y')}",
    )


@app.route("/view_payments/<int:class_id>")
def view_payments(class_id):
    # Get the start_date from the query string or default to the current date
    start_date_str = request.args.get("start_date", None)
    if start_date_str:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    else:
        start_date = datetime.today().replace(day=1)  # First day of the current month

    # Calculate the end_date (3 months from start_date)
    end_date = start_date + timedelta(days=90)  # Approx. 3 months

    # Fetch class data and payments filtered by date range
    class_data = model.get_class(class_id)
    payments = model.get_class_payments_in_date_range(class_id, start_date, end_date)

    # Extract unique names and dates within the range
    names = sorted(set(item["full_name"] for item in payments))
    dates = sorted(set(item["date"] for item in payments))

    # Build a lookup dictionary for quick access
    lookup = {(item["full_name"], item["date"]): item["paid"] for item in payments}

    student_payer = {
        item["full_name"]: {item["payer_name"], item["payer_phone_number"]}
        for item in payments
    }

    # Calculate previous and next start dates
    prev_start_date = (start_date - timedelta(days=90)).strftime("%Y-%m-%d")
    next_start_date = (start_date + timedelta(days=90)).strftime("%Y-%m-%d")

    return render_template(
        "view_payments.html",
        names=names,
        dates=dates,
        lookup=lookup,
        student_payer=student_payer,
        class_data=class_data,
        prev_start_date=prev_start_date,
        next_start_date=next_start_date,
        current_period=f"{start_date.strftime('%b %Y')} - {end_date.strftime('%b %Y')}",
    )


# Route to handle file upload
@app.route("/load_attendance", methods=["POST", "GET"])
def load_attendance():
    classes = model.get_classes_with_fields()
    if request.method != "POST":
        return render_template("load_attendance.html", classes=classes)

    file = request.files["file"]

    class_id = int(request.form["class_id"])
    fields_items = [
        f for f in (c["fields"] for c in classes if c["class_id"] == class_id)
    ][0]
    fields = {
        request.form["field_" + str(f["class_field_id"])]: f["class_field_id"]
        for f in fields_items
    }
    fields[request.form["attendance"]] = "attendance"
    print(fields)

    # try:
    data = process_file(file, fields)
    model.load_attendance_data(class_id, data)

    # Return data or render it in a template
    return f"Data from the first row: {data[0]}"

    # except Exception as e:
    #     return f"Error opening file: {e}"


def process_file(file, fields_names):
    file_stream = io.BytesIO(file.read())
    ws = openpyxl.load_workbook(file_stream).active

    def find_pattern_and_breakoff(dates, labels):
        # Step 1: Find the length of the repeating pattern in the labels.
        def find_repeating_pattern(lst):
            for length in range(1, len(lst) // 2 + 1):
                # Check if a pattern of length `length` repeats throughout the list
                pattern = lst[:length]
                repeated_pattern = pattern * (len(lst) // length)
                if repeated_pattern == lst[: len(repeated_pattern)]:
                    return length
            return len(
                lst
            )  # In case there's no repeating pattern, return the length of the list

        # Step 2: Find the break-off point where dates become all `None`
        def find_breakoff(dates, pattern_length):
            for i in range(0, len(dates), pattern_length):
                if dates[i] is None:
                    # Check if it's followed by a sequence of `None` values
                    if all(val is None for val in dates[i:]):
                        return i
            return len(
                dates
            )  # If no None sequence found, return the length of the list

        # Find the repeating pattern length in the labels
        repeating_pattern_length = find_repeating_pattern(labels)

        # Find the break-off point in the dates
        breakoff_point = find_breakoff(dates, repeating_pattern_length)

        return repeating_pattern_length, breakoff_point

    dates = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)][2:]
    fields = [ws.cell(row=2, column=i).value for i in range(1, ws.max_column + 1)][2:]

    num_fileds, num_days = find_pattern_and_breakoff(dates, fields)
    if num_fileds != len(fields_names):
        print(num_fileds, len(fields_names))
        raise Exception("INCORRECT NUMBER OF FIELDS")

    names = [ws.cell(row=i, column=2).value for i in range(3, ws.max_row + 1)]

    breakoff_index = next(
        (index for index, value in enumerate(names) if value is None), len(names)
    )

    attendance = []
    rows = list(
        ws.iter_rows(
            min_row=3,
            max_row=breakoff_index + 2,
            min_col=3,
            max_col=num_fileds * num_days + 2,
        )
    )

    field_names = [fields[i] for i in range(num_fileds)]
    # Loop over days
    attendance = [
        {
            "date": dates[day],
            "records": [
                {
                    "full_name": names[idx],
                    "fields": [
                        {
                            "field_id": fields_names[f],
                            "field_value": row[day + i].value,
                        }
                        for i, f in enumerate(field_names)
                    ],
                }
                for idx, row in enumerate(rows)
            ],
        }
        for day in range(0, num_days, num_fileds)
    ]

    return attendance


app.run(debug=True, host="0.0.0.0")

x = [
    {
        "date": datetime.datetime(2025, 1, 4, 0, 0),
        "records": [
            {
                "full_name": "yazeed",
                "fields": [
                    {"field_id": "attendance", "field_value": "yes"},
                    {"field_id": 1, "field_value": "done"},
                    {"field_id": 2, "field_value": "bad"},
                ],
            },
            {
                "full_name": "amin",
                "fields": [
                    {"field_id": "attendance", "field_value": "yes"},
                    {"field_id": 1, "field_value": "not done"},
                    {"field_id": 2, "field_value": "bad"},
                ],
            },
        ],
    },
]
